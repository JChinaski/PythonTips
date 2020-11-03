import csv
print("HOLAA")


from  openpyxl import load_workbook


def contar_caracter (caracter, x):
    """ La funcion contarA(x) cuenta cuántas
        letras "A" aparecen en la cadena x ."""
    contador = 0
    for letra in x:
        if letra == caracter:
            contador = contador + 1
            print(lista.find(caracter))
    return(contador)

def encuentra(cadena, carac, inicio=0):
    indice = inicio
    lista_indices = []
    cnt = inicio
    while indice < len(cadena):
        if cadena[indice] == carac:
            lista_indices[cnt] = indice
            cnt = cnt + 1
        indice += 1
    return lista_indices

"""
Estandarizar formato input o generar funcion para leer y guardar la lista si es q el dato trae basura

"""

wb = load_workbook('NUEVO_COMBI.xlsx')
print(wb.sheetnames) #['Sheet', 'Hoja1']

ir_a_hoja = wb['Hoja1']
wb.active = ir_a_hoja

print(f'Hoja activa: {wb.active.title}') #Hoja activa: Hoja1
print(wb.sheetnames) #['Sheet', 'Hoja1']
wb.save('NUEVO_COMBI.xlsx')
print(wb.sheetnames) #['Sheet', 'Hoja1']

#Acceder a una celda
hoja_activa = wb.active
a1 = hoja_activa["A1"]
print(a1.value) #None

hoja_activa["A1"] = """["uno", "dos", "tres","cuatro","cinco", "seis"]"""
a1 = hoja_activa["A1"]
print(wb.sheetnames) #['Sheet', 'Hoja1']
print(a1.value) #122333444455555 se puede borrar
print(wb.sheetnames) #['Sheet', 'Hoja1']

wb.save('NUEVO_COMBI.xlsx')

print(wb.sheetnames) #['Sheet', 'Hoja1']
print(a1.value) #122333444455555 se puede borrar

lista = a1.value
print ("tamano" , len(lista))
print (lista)
print (lista[3])


cadena = "Hola Mundo"

print(cadena.find("Hola"))        # Devuelve posición 0
print(cadena.find("Mundo"))       # Devuelve posición 5

print("cantidad registros lista")
print ( len(lista) )  

#print(lista.find('"'))        # Devuelve posición 0
#print(lista.find("cuatro"))       # Devuelve posición 5


#print (contar_caracter (",", lista))

#print (encuentra (lista, ","))
#print (encuentra (lista, '"'))

indice_comillas = encuentra (lista, ',')

print (indice_comillas)