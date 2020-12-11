import csv

from  openpyxl import load_workbook

print("TEST VERSION 000")


def respuestas ( resp , fil_col):
    ir_a_hoja = doc_excel['Respuestas']
    doc_excel.active = ir_a_hoja
    hoja_activa = doc_excel.active
    hoja_activa[fil_col] = resp
    ir_a_hoja = doc_excel['Preguntas']
    doc_excel.active = ir_a_hoja


def resultado():
    ir_a_hoja = doc_excel['Respuestas']
    doc_excel.active = ir_a_hoja
    hoja_activa = doc_excel.active
    contenido_Celda = hoja_activa['C66']
    return contenido_Celda.value

"""
Estandarizar formato input o generar funcion para leer y guardar la lista si es q el dato trae basura

"""

doc_excel = load_workbook('Test_1.xlsx') # CARGAMOS EXCEL
#print(doc_excel.sheetnames) #['Sheet', 'Hoja1'] -- muestra las pestañas del doc

ir_a_hoja = doc_excel['Preguntas'] #VAMOS A LA PESTAÑA EN Q SE ENCUENTRAN LAS PREGUNTAS
doc_excel.active = ir_a_hoja
hoja_activa = doc_excel.active

#print (ir_a_hoja) # ESTO ES PARA VER EN Q HOJA SE ESTA PARADO
#print(f'Hoja/Pestana activa: {doc_excel.active.title}') #Hoja activa: Hoja1 -- Muestra la pestaña activa "print (doc_excel.active)"

"""
#Acceder a una celda
i=1
for elemento in hoja_activa:
    fila_columna = hoja_activa["A" + f'{i}']
    print(fila_columna.value) #None
    #print(elemento)
    i=i+1

"""

# RECORRER EXCEL

i=1
Letra_Columa = ['B','C','D','E','F']

# RECORRER FILAS
for elemento in hoja_activa:
    fila_columna = hoja_activa["A" + f'{i}'] # SELECCIONAR CELDA -- EJ: A1, B1, C1 .... HASTA Q NO ENCUENTRA NADA
    if fila_columna.value != None:
        print(' ')
        print('Pregunta ' + f'{i}' + ' :')
        print(' ')
        print(fila_columna.value) # IMPRIME LO QUE CONTIENE LA CELDA DEFINIDA EN LAS LINEAS ANTERIORES
         # RECORRER COLUMNAS
        for e in Letra_Columa:
            indice_columna = e + f'{i}'
            #print(indice_columna) #N para ver que columna esta activa
            valor_columna = hoja_activa[indice_columna]
            if valor_columna.value != None:
                print(valor_columna.value) # imprime lo contenido en la celda
        

        # OBTENER RESPUESTA DE USUARIO
        print('Respuesta : ', end='')
        respuesta = input()
        respuestas ( respuesta.upper() , "A" + f'{i}') # FUNCION PARA GUARDAR RESPUESTA EN UNA PESTAÑA ESPECIFICA DEL DOCUMENTO
    
        i=i+1

doc_excel.save('Test_1.xlsx')

res= resultado()
print(res)



