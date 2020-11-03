## https://j2logo.com/python/crear-archivo-excel-en-python-con-openpyxl/
#Version 1: preguntas y respuestas en soco , directamente de la consola
#Version 2: preguntas y respuesta en el excel
#Version 3: pregutnas y respuestas de formas aleatorias

import csv

print("HOLAA")


#crea doc excel con una pestaña
import openpyxl
#wb = openpyxl.Workbook()

from  openpyxl import load_workbook
wb = load_workbook('productos_copia.xlsx')
print(wb.sheetnames) #['Sheet', 'Hoja1']

ir_a_hoja = wb['Hoja1']
wb.active = ir_a_hoja
print(f'Hoja activa: {wb.active.title}') #Hoja activa: Hoja1

wb.save('productos_copia.xlsx')

#Acceder a una celda
hoja_activa = wb.active
a1 = hoja_activa["A1"]
print(a1.value) #None

hoja_activa["A1"] = "¿Cual de los siguientes términos hace referencia a la potencia para aumentar o reducir con facilidad los recursos informáticos?"
a1 = hoja_activa["A1"]
#print(a1.value) #122333444455555 se puede borrar


#pregunta1

Respuesta1 = input()
print(f"Respuesta1, {Respuesta1}")

#hoja_activa = wb.active
#a2 = hoja_activa["A2"]
#print(a2.value) #None

hoja_activa["B1"] = Respuesta1
b1 = hoja_activa["B1"]
print(b1.value) #122333444455555 -nombre


wb.save('productos_copia.xlsx')

"""

hoja = wb['Hoja1']
wb.active = hoja
print(f'Hoja activa: {wb.active.title}') #Hoja activa: Hoja1

wb.save('productos_copia.xlsx')

#Acceder a una celda
wb = openpyxl.Workbook()
hoja = wb.active
a1 = hoja["A1"]
print(a1.value) #None

hoja["A1"] = 122333444455555
a1 = hoja["A1"]
print(a1.value) #122333444455555

wb.save('productos_copia.xlsx')

print("¿Cómo se llama? ", end="")
nombre = input()
print(f"Me alegro de conocerle, {nombre}")

"""