## https://j2logo.com/python/crear-archivo-excel-en-python-con-openpyxl/
import csv

print("HOLAA")


#crea doc excel con una pestaña
import openpyxl
wb = openpyxl.Workbook()

from openpyxl import load_workbook
wb = load_workbook('productos_copia.xlsx')
print(wb.sheetnames)

hoja = wb.active
print(f'Hoja activa: {hoja.title}')
hoja.title = "NuevaPestana"
print(f'Hoja activa: {wb.active.title}')

"""
# Añade la hoja 'Hoja' al final (por defecto)
>>> hoja1 = wb.create_sheet("Hoja")
# Añade la hoja 'Hoja' en la primera posición. Como el nombre
# 'Hoja' ya existe, le añade el número 1 al final del nombre
>>> hoja2 = wb.create_sheet("Hoja", 0)
# Añade la hoja 'Otra hoja' en la posición 1
>>> wb.create_sheet(index=1, title="Otra hoja")
# Muestra los nombres de las hojas
>>> print(wb.sheetnames)
['Hoja1', 'Otra hoja', 'Valores', 'Hoja']
"""
print(wb.sheetnames)
"""
#También es posible crear una copia de una hoja con el método copy_worksheet():
origen = wb.active
nueva = wb.copy_worksheet(origen)
print(wb.sheetnames)
"""

wb.save('productos.xlsx')

"""#Acceder a una hoja
>>> hoja = wb.active  # Es la hoja que está en el índice 0
>>> print(f'Hoja activa: {hoja.title}')
Hoja activa: Hoja1
>>> hoja = wb['Otra hoja']
>>> wb.active = hoja
>>> print(f'Hoja activa: {wb.active.title}')
Hoja activa: Otra hoja
"""

hoja = wb['NuevaPestana']  #['Hoja1']
wb.active = hoja
print(f'Hoja activa: {wb.active.title}')

#Acceder a una celda
wb = openpyxl.Workbook()
hoja = wb.active
a1 = hoja["A1"]
print(a1.value)


"""
#Escribir valores en una celda
# 1.- Asignando el valor directamente a la celda
>>> hoja["A1"] = 10
>>> a1 = hoja["A1"]
>>> print(a1.value)
10
# 2.- Usando la notación fila, columna con el argumento value
>>> b1 = hoja.cell(row=1, column=2, value=20)
>>> print(b1.value)
20
# 3.- Actualizando la propiedad value de una celda
>>> c1 = hoja.cell(row=1, column=3)
>>> c1.value = 30
>>> print(c1.value)
30
"""

hoja["A1"] = 57894
a1 = hoja["A1"]
print(a1.value)

hoja["B3"] = 'tres'
B3 = hoja["B3"]
print(a1.value)

wb.save('productos.xlsx')
# Añade la hoja 'Hoja' al final (por defecto)
hoja1 = wb.create_sheet("Hoja1") #CREAR UNA PESTAÑA NUEVA


wb.save('productos.xlsx')


"""
with open('./Documentos/TEST_1/names.csv', 'r') as csv_file:
    csv_reader = csv.DictReader(csv_file)

    with open('./Documentos/TEST_1/new_names.csv', 'w') as new_file:
        fieldnames = ['first_name', 'last_name']

        csv_writer = csv.DictWriter(new_file, fieldnames=fieldnames, delimiter='\t')

        csv_writer.writeheader()

        for line in csv_reader:
            del line['email']
            csv_writer.writerow(line)
"""