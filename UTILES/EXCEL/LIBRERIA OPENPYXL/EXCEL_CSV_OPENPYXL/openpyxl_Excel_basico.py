## https://j2logo.com/python/crear-archivo-excel-en-python-con-openpyxl/

import csv

from  openpyxl import load_workbook

"""
SE DEBE Estandarizar formato input o generar funcion para leer y guardar la lista si es q el dato trae basura

"""

wb = load_workbook('Template_ejem_openpyxl.xlsx')
print("Muestra las pestañas del documento", wb.sheetnames) #['Sheet', 'Hoja1']

ir_a_hoja = wb['Hoja1']
wb.active = ir_a_hoja

print(f'Hoja activa: {wb.active.title}') #Hoja activa: Hoja1 #Muestra la pestaña activa
print(wb.sheetnames) #['Sheet', 'Hoja1'] #Muestra la pestaña activa
wb.save('Template_ejem_openpyxl.xlsx') #SALVAR DOCUMENTO
print("Documento Salvado")
print(wb.sheetnames) #['Sheet', 'Hoja1']

#Acceder a una celda
hoja_activa = wb.active
a1 = hoja_activa["A1"]
print(a1.value) # MUESTRA LO QUE CONTIENE LA CELDA ESPECIFICADA EN LA LINEA ANTERIOR

#AGREGAR DATO A UNA CELDA
hoja_activa["A1"] = """["uno", "dos", "tres","cuatro","cinco", "seis"]""" #ESTO ES LO QUE SE QUIERE INGRESAR A LA CELDA
a1 = hoja_activa["A1"] # aCA SE INGRESA EL VALOR ANTERIOR
print(a1.value) 


wb.save('Template_ejem_openpyxl.xlsx') #SALVAR DOCUMENTO
print("Documento Salvado")

"""

# TRABAJAR EL DATO DE UNA CELDA COMO UN String

lista = a1.value # Rescatar dato de celda
print ("tamano" , len(lista)) # Obtener el tamaño de la lista
print (lista) # mostrar lista
print (lista[3]) # Selecionar mediante indice de la lista
print("cantidad registros lista")
print ( len(lista) )
print(lista.find('"'))        # Devuelve posición 0
print(lista.find("cuatro"))       # Devuelve posición 5


#Ejemplo de como buscar en una lista
cadena = "Hola Mundo"

print(cadena.find("Hola"))        # Devuelve posición 0
print(cadena.find("Mundo"))       # Devuelve posición 5
print ( len(cadena) ) 

"""